import json

from questions import QUESTIONS, OPEN_QUESTIONS

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter


# ======================================================
# Стили
# ======================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="123A72"
)

SECTION_FILL = PatternFill(
    fill_type="solid",
    fgColor="EAF2FD"
)

WHITE_FONT = Font(
    bold=True,
    color="FFFFFF",
    size=12
)

TITLE_FONT = Font(
    bold=True,
    size=16,
    color="123A72"
)

BOLD_FONT = Font(
    bold=True,
    size=11
)

NORMAL_FONT = Font(
    size=11
)

BIG_FONT = Font(
    bold=True,
    size=22,
    color="123A72"
)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

LEFT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True
)

TOP_LEFT = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True
)

THIN = Side(
    style="thin",
    color="BFBFBF"
)

BORDER = Border(
    left=THIN,
    right=THIN,
    top=THIN,
    bottom=THIN
)


# ======================================================
# Экспорт одной анкеты
# ======================================================

def export_one_survey(data):

    wb = Workbook()

    ws = wb.active

    ws.title = "Анкета"

    # --------------------------------------------------
    # Шапка
    # --------------------------------------------------

    ws.merge_cells("A1:C1")

    title = ws["A1"]

    title.value = (
        "ТОО «CT Assembly»\n"
        "Анкета оценки производственной практики"
    )

    title.font = TITLE_FONT
    title.alignment = CENTER

    ws.row_dimensions[1].height = 60

    # пустая строка

    ws.append([])

    # --------------------------------------------------
    # Заголовок карточки
    # --------------------------------------------------

    ws.append(["Поле", "Значение"])

    header_row = ws.max_row

    for cell in (ws[f"A{header_row}"], ws[f"B{header_row}"]):
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    from openpyxl.styles import PatternFill, Border

    ws[f"C{header_row}"].fill = PatternFill(fill_type=None)
    ws[f"C{header_row}"].border = Border()

    # --------------------------------------------------
    # Карточка студента
    # --------------------------------------------------

    student = [

        ["№ анкеты", data[0]],

        ["Дата", data[1]],

        ["Студент", data[2]],

        ["Специальность", data[3]],

        ["Курс", data[4]],

        ["Предприятие", data[5]],

        ["Наставник", data[6]]

    ]

    for row in student:

        ws.append(row)

        r = ws.max_row

        ws[f"A{r}"].font = BOLD_FONT
        ws[f"A{r}"].fill = SECTION_FILL
        ws[f"A{r}"].alignment = LEFT

        ws[f"B{r}"].font = NORMAL_FONT
        ws[f"B{r}"].alignment = LEFT

        ws[f"A{r}"].border = BORDER
        ws[f"B{r}"].border = BORDER

    answers = json.loads(data[7])
        # --------------------------------------------------
    # Таблица компетенций
    # --------------------------------------------------

    ws.append([])

    ws.append([
        "№",
        "Критерий оценки",
        "Оценка"
    ])

    score_header_row = ws.max_row

    for cell in ws[score_header_row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for i, answer in enumerate(answers):

        ws.append([
            i + 1,
            QUESTIONS[i],
            answer
        ])

        r = ws.max_row

        # Номер
        ws[f"A{r}"].alignment = CENTER
        ws[f"A{r}"].border = BORDER

        # Критерий
        ws[f"B{r}"].alignment = LEFT
        ws[f"B{r}"].border = BORDER

        # Оценка
        ws[f"C{r}"].alignment = CENTER
        ws[f"C{r}"].border = BORDER

        # Чередование цветов строк
        if i % 2 == 0:

            for col in ("A", "B", "C"):

                ws[f"{col}{r}"].fill = SECTION_FILL

    score_last_row = ws.max_row

    # --------------------------------------------------
    # Средний балл
    # --------------------------------------------------

    ws.append([])

    ws.append(["СРЕДНИЙ БАЛЛ"])

    r = ws.max_row

    ws.merge_cells(
        start_row=r,
        start_column=1,
        end_row=r,
        end_column=3
    )

    cell = ws.cell(r, 1)

    cell.font = WHITE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

    average = float(data[8])

    ws.append([f"{average:.2f}"])

    r = ws.max_row

    ws.merge_cells(
        start_row=r,
        start_column=1,
        end_row=r,
        end_column=3
    )

    cell = ws.cell(r, 1)

    cell.font = BIG_FONT
    cell.alignment = CENTER
    cell.border = BORDER
        # --------------------------------------------------
    # Комментарии наставника
    # --------------------------------------------------

    comments = [

        (OPEN_QUESTIONS[0], data[9]),

        (OPEN_QUESTIONS[1], data[10]),

        (OPEN_QUESTIONS[2], data[11])

    ]

    for title, text in comments:

        ws.append([])

        # Заголовок блока

        ws.append([title])

        r = ws.max_row

        ws.merge_cells(
            start_row=r,
            start_column=1,
            end_row=r,
            end_column=3
        )

        cell = ws.cell(r, 1)

        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT
        cell.border = BORDER

        # Текст комментария

        ws.append([text if text else "-"])

        r = ws.max_row

        ws.merge_cells(
            start_row=r,
            start_column=1,
            end_row=r,
            end_column=3
        )

        cell = ws.cell(r, 1)

        cell.font = NORMAL_FONT
        cell.alignment = TOP_LEFT
        cell.border = BORDER

        ws.row_dimensions[r].height = 55
            # --------------------------------------------------
    # Настройка ширины столбцов
    # --------------------------------------------------

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 85
    ws.column_dimensions["C"].width = 14

    # --------------------------------------------------
    # Высота строк
    # --------------------------------------------------

    # Заголовок документа
    ws.row_dimensions[1].height = 60

    # Строки карточки студента
    for row in range(4, 11):
        ws.row_dimensions[row].height = 24

    # Строки с критериями оценки
    for row in range(score_header_row + 1, score_last_row + 1):
        ws.row_dimensions[row].height = 38

    # --------------------------------------------------
    # Автофильтр
    # --------------------------------------------------

    ws.auto_filter.ref = (
        f"A{score_header_row}:C{score_last_row}"
    )

    # --------------------------------------------------
    # Закрепить только первую строку
    # --------------------------------------------------

    ws.freeze_panes = "A2"

    # --------------------------------------------------
    # Скрыть сетку
    # --------------------------------------------------

    ws.sheet_view.showGridLines = False

    # --------------------------------------------------
    # Настройки страницы
    # --------------------------------------------------

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    ws.print_options.gridLines = False

    # --------------------------------------------------
    # Область печати
    # --------------------------------------------------

    ws.print_area = f"A1:C{ws.max_row}"

    # --------------------------------------------------
    # Повторять первую строку при печати
    # --------------------------------------------------

    ws.print_title_rows = "1:1"

    # --------------------------------------------------
    # Центрирование по ширине листа
    # --------------------------------------------------

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # --------------------------------------------------
    # Сохранение
    # --------------------------------------------------

    filename = f"survey_{data[0]}.xlsx"

    wb.save(filename)

    return filename