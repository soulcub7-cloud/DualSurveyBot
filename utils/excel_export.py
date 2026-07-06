import json

from questions import QUESTIONS

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter

from openpyxl.formatting.rule import ColorScaleRule

# --------------------------------------------------
# Стили
# --------------------------------------------------

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="123A72"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

BOLD_FONT = Font(
    bold=True
)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

LEFT = Alignment(
    vertical="top",
    wrap_text=True
)

THIN = Side(
    style="thin",
    color="C8C8C8"
)

BORDER = Border(
    left=THIN,
    right=THIN,
    top=THIN,
    bottom=THIN
)

def export_surveys_to_excel(data):

    wb = Workbook()

    ws = wb.active

    ws.title = "Анкеты"

    headers = [

        "№",

        "Дата",

        "Студент",

        "Специальность",

        "Курс",

        "Предприятие",

        "Наставник"

    ]

    headers.extend(QUESTIONS)

    headers.extend([

        "Средний балл",

        "Лучшие качества",

        "Что необходимо улучшить",

        "Рекомендации"

    ])

    ws.append(headers)

    for cell in ws[1]:

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = CENTER

        cell.border = BORDER

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # --------------------------------------------------
    # Заполнение таблицы
    # --------------------------------------------------

    for survey in data:

        answers = json.loads(survey[7])

        row = [

            survey[0],
            survey[1],
            survey[2],
            survey[3],
            survey[4],
            survey[5],
            survey[6]

        ]

        row.extend(answers)

        row.extend([

            float(survey[8]),

            survey[9],

            survey[10],

            survey[11]

        ])

        ws.append(row)
    # --------------------------------------------------
    # Оформление таблицы
    # --------------------------------------------------

    for row in ws.iter_rows(min_row=2):

        for cell in row:

            cell.border = BORDER

            if cell.column <= 23:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
    
    # --------------------------------------------------
    # Цветовая шкала среднего балла
    # --------------------------------------------------

    avg_column = get_column_letter(23)

    ws.conditional_formatting.add(

        f"{avg_column}2:{avg_column}{ws.max_row}",

        ColorScaleRule(

            start_type="num",
            start_value=2,
            start_color="F8696B",

            mid_type="num",
            mid_value=3.5,
            mid_color="FFEB84",

            end_type="num",
            end_value=5,
            end_color="63BE7B"

        )

    )

    # --------------------------------------------------
    # Автоматическая ширина
    # --------------------------------------------------

    for column in ws.columns:

        length = 0

        letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > length:
                    length = len(str(cell.value))

            except Exception:
                pass

        width = min(max(length + 2, 12), 45)

        ws.column_dimensions[letter].width = width
        
    filename = "surveys.xlsx"

    wb.save(filename)

    return filename