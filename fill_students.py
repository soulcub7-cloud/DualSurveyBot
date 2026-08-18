from pathlib import Path

from openpyxl import load_workbook

from database import (
    add_enterprise,
    add_student,
    deactivate_missing_enterprises,
    deactivate_missing_students,
)


FILE_NAME = Path(__file__).resolve().parent / "students.xlsx"


def optional_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def optional_int(value):
    if value is None or str(value).strip() == "":
        return None
    return int(value)


def sync_reference_data():
    workbook = load_workbook(FILE_NAME, data_only=True)
    students_sheet = (
        workbook["Студенты"]
        if "Студенты" in workbook.sheetnames
        else workbook.active
    )

    added = 0
    updated = 0
    actual_students = []

    for row in students_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        fio = str(row[0]).strip()
        stream = int(row[1])
        speciality = optional_text(row[2] if len(row) > 2 else None)
        course = optional_int(row[3] if len(row) > 3 else None)

        actual_students.append(fio)

        if add_student(fio, stream, speciality, course):
            added += 1
        else:
            updated += 1

    deactivated_students = deactivate_missing_students(actual_students)

    actual_enterprises = []
    if "Предприятия" in workbook.sheetnames:
        enterprises_sheet = workbook["Предприятия"]

        for row in enterprises_sheet.iter_rows(min_row=2, values_only=True):
            name = optional_text(row[0])
            if not name:
                continue

            actual_enterprises.append(name)
            add_enterprise(name)

        deactivated_enterprises = deactivate_missing_enterprises(
            actual_enterprises
        )
    else:
        deactivated_enterprises = 0

    print(
        "Справочники синхронизированы: "
        f"добавлено студентов {added}, обновлено {updated}, "
        f"скрыто {deactivated_students}; "
        f"предприятий {len(actual_enterprises)}, "
        f"скрыто {deactivated_enterprises}."
    )

    return {
        "added_students": added,
        "updated_students": updated,
        "deactivated_students": deactivated_students,
        "enterprises": len(actual_enterprises),
        "deactivated_enterprises": deactivated_enterprises,
    }


if __name__ == "__main__":
    sync_reference_data()
